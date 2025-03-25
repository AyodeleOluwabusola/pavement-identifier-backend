import json
import logging
import os
import time
import threading
from pathlib import Path
from typing import Dict, Any, List
from threading import Event, Thread, Lock
import pika
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from app.core.config import settings
from app.core.logger import setup_logging
from app.services.rabbitmq_service import get_rabbitmq_connection

logger = setup_logging(__name__)


class HighPerformanceWriter:
    def __init__(self, output_path: Path, batch_size: int, flush_interval: int):
        self.output_path = output_path
        self.batch_size = batch_size
        self.flush_interval = flush_interval
        self.buffer = []
        self.lock = Lock()
        self.stop_event = threading.Event()
        self.thread = threading.Thread(target=self._flush_buffer)
        self.thread.start()

    def write(self, data: Dict[str, Any]):
        with self.lock:
            self.buffer.append(data)
            if len(self.buffer) >= self.batch_size:
                self._flush()

    def _flush(self):
        if self.buffer:
            self._save_to_excel(self.buffer)
            self.buffer.clear()

    def _save_to_excel(self, data: List[Dict[str, Any]]):
        try:
            workbook = load_workbook(self.output_path)
        except InvalidFileException:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Image Processing Results"
            sheet.append(
                ["File Name", "Status", "Predicted Class", "Confidence"])

        sheet = workbook.active
        for item in data:
            sheet.append([
                item.get('file_name', 'Unknown'),
                item.get('status', 'Error'),
                item.get('predicted_class', 'Unknown'),
                item.get('confidence', 0.0)
            ])

        workbook.save(self.output_path)

    def _flush_buffer(self):
        while not self.stop_event.is_set():
            time.sleep(self.flush_interval)
            with self.lock:
                self._flush()

    def stop(self):
        self.stop_event.set()
        self.thread.join()
        with self.lock:
            self._flush()
            self._save_to_excel(self.buffer)
            self.buffer.clear()


class ExcelWriterService:
    def __init__(self):
        self.writer = HighPerformanceWriter(
            output_path=Path(settings.EXCEL_RESULTS_PATH),
            batch_size=1000,  # Adjust based on your needs
            flush_interval=30  # Seconds
        )

    def write_result(self, data: Dict[str, Any]) -> bool:
        try:
            self.writer.write(data)
            return True
        except Exception as e:
            logger.error(f"Failed to write data: {e}")
            return False

    def cleanup(self):
        self.writer.stop()


class ExcelWriterManager:
    """Manages multiple Excel writer threads"""

    def __init__(self, num_consumers: int = 3):
        self.num_consumers = num_consumers
        self.writer_threads: List[Thread] = []
        self._ready = Event()
        self._stopped = Event()
        self.shared_writer = ExcelWriterService()  # Share a single writer instance

    def _run_writer_thread(self):
        """Function to run in each writer thread"""
        thread_logger = setup_logging(
            f"excel_writer_thread_{threading.get_ident()}")
        thread_logger.info(
            f"Starting Excel writer thread {threading.get_ident()}")

        try:
            def callback(ch, method, properties, body):
                try:
                    thread_logger.info("Received message from results queue")
                    data = json.loads(body)
                    thread_logger.info(
                        f"Processing message for file: {data.get('file_name')}")

                    success = self.shared_writer.write_result(data)

                    if success:
                        ch.basic_ack(delivery_tag=method.delivery_tag)
                        thread_logger.info(
                            f"Successfully processed message for {data.get('file_name')}")
                    else:
                        ch.basic_reject(
                            delivery_tag=method.delivery_tag, requeue=True)
                        thread_logger.warning(
                            f"Failed to process message for {data.get('file_name')}, requeueing")

                except json.JSONDecodeError as e:
                    thread_logger.error(f"Invalid JSON in message: {e}")
                    ch.basic_reject(
                        delivery_tag=method.delivery_tag, requeue=False)
                except Exception as e:
                    thread_logger.error(f"Error processing message: {e}")
                    ch.basic_reject(
                        delivery_tag=method.delivery_tag, requeue=True)

            # Set up consumer connection for this thread
            connection = pika.BlockingConnection(
                pika.ConnectionParameters(
                    host=settings.RABBITMQ_HOST,
                    port=settings.RABBITMQ_PORT,
                    credentials=pika.PlainCredentials(
                        settings.RABBITMQ_USER,
                        settings.RABBITMQ_PASSWORD
                    ),
                    virtual_host=settings.RABBITMQ_VHOST
                )
            )
            channel = connection.channel()

            # Set up consumer
            channel.basic_qos(prefetch_count=1)
            channel.basic_consume(
                queue=settings.RESULTS_QUEUE_NAME,
                on_message_callback=callback
            )

            thread_logger.info(
                "Excel writer thread is ready to consume messages")

            # Start consuming until stopped
            while not self._stopped.is_set():
                try:
                    # Process messages with timeout
                    connection.process_data_events(time_limit=1)
                except Exception as e:
                    if not self._stopped.is_set():
                        thread_logger.error(f"Error processing messages: {e}")
                        time.sleep(1)  # Wait before reconnecting

            # Clean up
            try:
                channel.close()
                connection.close()
            except Exception as e:
                thread_logger.error(f"Error closing connections: {e}")

        except Exception as e:
            thread_logger.error(f"Excel writer thread error: {e}")
            if not self._stopped.is_set():
                self._ready.clear()

    def start(self):
        """Start all Excel writer threads"""
        logger.info(f"Starting {self.num_consumers} Excel writer threads")

        try:
            for i in range(self.num_consumers):
                if self._stopped.is_set():
                    break

                thread = Thread(
                    target=self._run_writer_thread,
                    name=f"excel_writer_{i}",
                    daemon=True
                )
                thread.start()
                self.writer_threads.append(thread)
                logger.info(f"Started Excel writer thread {i+1}")

            if self.writer_threads and not self._stopped.is_set():
                self._ready.set()
                logger.info("Excel writer manager is ready")

        except Exception as e:
            logger.error(f"Error starting Excel Writer Manager: {e}")
            self.stop()
            raise

    def stop(self):
        """Stop all Excel writer threads"""
        logger.info("Stopping Excel writer manager")
        self._stopped.set()
        self._ready.clear()

        # Wait for threads to finish
        for thread in self.writer_threads:
            try:
                thread.join(timeout=5)  # Give threads 5 seconds to finish
            except Exception as e:
                logger.error(f"Error joining thread: {e}")

        self.writer_threads.clear()
        logger.info("Excel writer manager stopped")

    def is_ready(self) -> bool:
        """Check if the service is ready"""
        return self._ready.is_set() and any(thread.is_alive() for thread in self.writer_threads)


def run_excel_writer(num_consumers: int = 3) -> ExcelWriterManager:
    """Start the Excel writer manager with multiple consumers"""
    try:
        manager = ExcelWriterManager(num_consumers)
        manager.start()
        logger.info(
            f"Started Excel writer manager with {num_consumers} consumers")
        return manager
    except Exception as e:
        logger.error(f"Critical error in Excel writer manager: {e}")
        raise
